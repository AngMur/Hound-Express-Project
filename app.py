from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import zipfile
from io import BytesIO
import pandas as pd
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')


# Rutas
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

EXCELTEMPALTES_FOLDER = 'excel_templates'
app.config['EXCELTEMPLATES_FOLDER'] = EXCELTEMPALTES_FOLDER

# Verificar que la ruta para subir archivos existe
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/upload/<string:option>', methods=['POST'])
def upload_file(option):
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    # Guardar el archivo
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    # Llamar a la función de procesamiento
    output_file =  generar_facturas(file_path) if option == "Facturas" else process_excel(file_path)

    if(option == "Facturas"):
        # Comprimir los archivos procesados en un archivo ZIP
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for file_path in output_file:
                zip_file.write(file_path, os.path.basename(file_path))

        zip_buffer.seek(0)  # Posicionar el puntero del archivo en el inicio

        return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name='facturas.zip')

    # return jsonify({'message': 'File uploaded and proccessed successfully', 'output_file': output_file})
    return send_file(output_file, as_attachment=True)


# Generador de separaciones---------------------------------------------------------------
def process_excel(filepath):
    # Procesar el archivo excel subido
    df = pd.read_excel(filepath)
    df_original = df

    # Convertir las columnas a tipos adecuados
    df['Tracking Number (HAWB)'] = df['Tracking Number (HAWB)'].astype(str)
    df['TOTAL QTY OF ITEMS IN PARCEL'] = df['TOTAL QTY OF ITEMS IN PARCEL'].astype(float)
    df['TOTAL DECLARED VALUE'] = df['TOTAL DECLARED VALUE'].astype(float)

    # Agregar la columna IVA
    df['IVA'] = df['TOTAL DECLARED VALUE'].apply(lambda x: 0.0 if x < 50.01 else (0.17 if 50.01 <= x <= 117.01 else 0.19))

    # Mapear los shippers
    ships = {"IMEX - Mattel One Shop": "JUGUETE", "FragranceNet.com": "PERFUME"}
    df['SHORT DESCRIPTION'] = df['SHIPPER'].map(lambda x: ships.get(x, ""))

    # Definir condiciones de filtrado
    condiciones_filtrado = (df['Tracking Number (HAWB)'].str.len() == 22) | (df['TOTAL QTY OF ITEMS IN PARCEL'] > 10) | (df['TOTAL DECLARED VALUE'] >= 500) | (df['PRODUCT DESCRIPTION'].str.contains(r'\bother\b', case=False))

    # Separar registros especiales y normales
    df_especiales = df[condiciones_filtrado]
    df_normal = df[~condiciones_filtrado].sort_values(by='TOTAL DECLARED VALUE')

    # Separar registros normales en menores y mayores
    df_menores = df_normal[df_normal['TOTAL DECLARED VALUE'] < 50.01].reset_index(drop=True)
    df_mayores = df_normal[df_normal['TOTAL DECLARED VALUE'] >= 50.01].reset_index(drop=True)

    # Crear bloques de registros mayores
    limite = 5000
    bloques = []
    sumatoria = 0.0
    inicio_b = 0
    for index, log in df_mayores.iterrows():
        actual_price = float(log['TOTAL DECLARED VALUE'])
        if (sumatoria + actual_price) <= limite:
            sumatoria += actual_price
        else:
            bloques.append((inicio_b, index))
            inicio_b = index
            sumatoria = actual_price
    bloques.append((inicio_b, len(df_mayores)))

    # Crear DataFrames de secciones
    agrupacion_dfs = {"MENORES": df_menores}
    for i, bloque in enumerate(bloques):
        identificador = f"MAYORES {i + 1}"
        agrupacion_dfs[identificador] = df_mayores.iloc[bloque[0]:bloque[1]]
    agrupacion_dfs['ESPECIALES'] = df_especiales

    # Inicializar df_final
    df_final = pd.DataFrame()

    # Agrupamos los bloques obtenidos
    def agrupar_bloque(titulo, bloque_df):
        df_title = pd.DataFrame({'GRUPO': [titulo]})
        return pd.concat([df_title, bloque_df], ignore_index=True)

    # Concatenar DataFrames en df_final
    for key, value in agrupacion_dfs.items():
        df_final = pd.concat([df_final, agrupar_bloque(key, value)], ignore_index=True)

    # Guardar el archivo filtrado
    output_filename = 'datos_filtrados.xlsx'
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    writer = pd.ExcelWriter(output_path, engine='xlsxwriter')

    # Escribir cada DataFrame en una hoja diferente del archivo de Excel
    df_final.to_excel(writer, sheet_name='Separados', index=False)
    df_especiales.to_excel(writer, sheet_name='Especiales', index=False)
    df_normal.to_excel(writer, sheet_name='Normales', index=False)
    df_mayores.to_excel(writer, sheet_name='Mayores', index=False)
    df_menores.to_excel(writer, sheet_name='Menores', index=False)
    df_original.to_excel(writer, sheet_name='Originales', index=False)


    # Estilos para las hojas (backgrounds)
    worksheet = writer.sheets['Separados']
    worksheet.set_tab_color('orange')
    worksheet = writer.sheets['Especiales']
    worksheet.set_tab_color('purple')
    worksheet = writer.sheets['Normales']
    worksheet.set_tab_color('red')
    worksheet = writer.sheets['Mayores']
    worksheet.set_tab_color('blue')
    worksheet = writer.sheets['Menores']
    worksheet.set_tab_color('yellow')

    # Encontrar los índices de los valores distintos de NaN en la columna 'GRUPOS'
    titulos = df_final[df_final['GRUPO'].notna()].index

    workbook  = writer.book
    worksheet = writer.sheets['Separados']

    # Definir un formato para la fila
    formato_fila = workbook.add_format({'bg_color': 'pink', 'bold': True, 'font_size': 26})

    # Ajustar el ancho de la fila y aplicar el formato a cada fila con titulo
    for i in titulos:
        worksheet.set_row(i + 1, 35, formato_fila)
        
    # Cerrar el escritor
    writer.close()

    return output_path

#Generador de facturas-------------------------------------------------------------
def generar_facturas(filepath):
    # Leer el archivo Excel subido y procesar los bloques
    df_filtrado = pd.read_excel(filepath)
    bloques = separar_bloques(df_filtrado)
    files = []
    # Generar facturas para cada bloque
    for indice, bloque in enumerate(bloques, start=0):
        if indice == 0:
            files.append(generar_documento("FACTURA MENOR", bloque, False))

        else:
            files.append(generar_documento(f"FACTURA MAYOR{indice}", bloque, True))
    
    return files


def generar_documento(titulo, bloque, mayor="False"):
    
    template_name = 'plantilla_mayor.xlsx' if mayor else 'plantilla_menor.xlsx'
    file_path = os.path.join(app.config['EXCELTEMPLATES_FOLDER'], template_name)
    # Cargar el archivo Excel existente
    wb = load_workbook(file_path)
    ws = wb.active
    ws.delete_rows(7 if mayor else 9)
    generar_info(bloque, ws, mayor)

    output_filename = f'{titulo}.xlsx'
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)

    wb.save(output_path)

    return output_path


def separar_bloques(df):
    # Obtener los indices de aquellas celdas que no están vacías en 'GRUPO'
    indices = df[df['GRUPO'].notna()].index
    bloques = []

    # Separar los bloques basados en los índices
    for i in range(len(indices) - 1):
        inicio = indices[i] + 1
        fin = indices[i + 1]
        bloques.append(df.iloc[inicio:fin])
    return bloques


def generar_info(bloque, ws, mayor="False"):
    guia = 1
    for indice in range(len(bloque)):
        registro = bloque.iloc[indice]
        if mayor:
            info = [
                guia,
                registro['Tracking Number (HAWB)'],
                registro['TOTAL QTY OF ITEMS IN PARCEL'],
                "Paquete",
                registro['TOTAL QTY OF ITEMS IN PARCEL'],
                "Pz",
                registro['SHORT DESCRIPTION'],
                registro['TOTAL DECLARED VALUE'],
                (registro['TOTAL DECLARED VALUE'] * registro['TOTAL QTY OF ITEMS IN PARCEL']),
                "USA",
                "HOUND EXPRESS"
            ]
        else:
            info = [
                guia,
                registro['Tracking Number (HAWB)'],
                1,
                "Paquete",
                1,
                registro['SHORT DESCRIPTION'],
                registro['TOTAL DECLARED VALUE'],
                "USA",
                "HOUND EXPRESS"
            ]
        aumento = 6 if mayor else 8
        ws.insert_rows(guia + aumento, amount=1)
        rellenar_fila(guia + aumento, info, ws, mayor)
        aplicar_estilos(guia + aumento, ws, mayor)
        guia += 1

    if mayor:
        ws[f'I{guia+aumento}'] = (bloque['TOTAL DECLARED VALUE'] * bloque['TOTAL QTY OF ITEMS IN PARCEL']).sum()
        ws[f'I{guia+aumento}'].number_format = '$#,##0.00'
        ws[f'E{guia+aumento}'] = bloque['TOTAL QTY OF ITEMS IN PARCEL'].sum()
    else:
        ws[f'G{guia+aumento}'] = bloque['TOTAL DECLARED VALUE'].sum()
        ws[f'G{guia+aumento}'].number_format = '$#,##0.00'


def rellenar_fila(fila_a_insertar, info, ws, mayor="False"):
    for idx, value in enumerate(info, start=1):
        ws.cell(row=fila_a_insertar, column=idx, value=value)

def aplicar_estilos(fila_a_formatear, ws, mayor="False"):
    ws.row_dimensions[fila_a_formatear].height = 17.5
    font_style = Font(name="Arial Narrow", size=13)
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[fila_a_formatear]:
        cell.font = font_style
        cell.border = border_style
        cell.alignment = alignment

    cell = ws[f"A{fila_a_formatear}"]
    cell.font = Font(bold=True)
    if mayor:
        cell = ws[f"K{fila_a_formatear}"]
        cell.font = Font(bold=True)
        cell = ws[f"H{fila_a_formatear}"]
        cell.number_format = '$#,##0.00'
        cell = ws[f"I{fila_a_formatear}"]
        cell.number_format = '$#,##0.00'
    else:
        cell = ws[f"I{fila_a_formatear}"]
        cell.font = Font(bold=True)
        cell = ws[f"G{fila_a_formatear}"]
        cell.number_format = '$#,##0.00'






if __name__ == "__main__":
    app.run(debug=True)
